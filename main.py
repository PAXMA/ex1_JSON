import json
import openpyxl
import traceback
import sys


class JSONToExcelParser:
    def __init__(self):
        self.wb = openpyxl.Workbook()  # Excel-файл
        self.is_first_ever_sheet = True  # Флаг первого пустого листа

    def save_as_excel(self, name):
        """
        Сохраняет текущий Excel-файл под указанными именем.

        :param str name: Имя файла.
        :return:
        """
        try:
            self.wb.save(f"{name}.xlsx")
            print(f"Excel-файл сохранён под именем {name}.")
        except:
            print(f"create_excel_file: {traceback.format_exc()}")

    def parse_json(self, json_file):
        """
        Парсит указанный JSON-файл и записывает полученные данные в лист в Excel.

        :param str json_file: Имя JSON-файла.
        :return:
        """
        try:
            with open(json_file, encoding="utf-8") as f:
                data = json.load(f)
                if not ("headers" in data and "values" in data):
                    print(f"{json_file} не является подходящим JSON-файлом.")
                    return
                list_for_excel = self.search_field_value(data["headers"], "QuickInfo")
                if not list_for_excel:
                    print(f"JSON-файлом {json_file} будет проигнорирован из-за возникшей ошибки.")
                    return
                values_list = self.search_field_value(data["values"], "Text")
                if not values_list:
                    print(f"JSON-файлом {json_file} будет проигнорирован из-за возникшей ошибки.")
                    return
                list_for_excel += values_list

                if not self.is_first_ever_sheet:
                    if json_file in self.wb.sheetnames:
                        print(f"JSON-файл {json_file} уже был обработан.")
                        return
                    self.wb.create_sheet(json_file)
                    self.wb.active = self.wb[json_file]
                else:
                    self.wb.active.title = json_file
                for chunk in list_for_excel:
                    self.wb.active.append(chunk)
                self.is_first_ever_sheet = False
                print(f"JSON-файл {json_file} успешно обработан.")
                return
        except FileNotFoundError:
            print(f"Файл c именем {json_file} не найден.")
            return
        except:
            print(f"parse_json: {traceback.format_exc()}")
            return

    def search_field_value(self, array, field_name):
        """
        Ищет значение указанного поля.
        Для сортировки значений по используются значения поля "MaxLength" (Порядок значений: 10 -> 7 -> 2 -> 9)

        :param list[dict] array: Список словарей для поиска.
        :param str field_name: Искомое поле.
        :return: Список значений.
        :rtype: list[str]
        """
        chunk = 4  # Кол-во значений в одной группе (кол-во столбцов на листе)
        raw_list = []  # Список найденных значений полей
        # TODO: можно изобрести более универсальный поиск, проходя по всем элементам списка и определяя их тип.
        for each_dict in array:
            for key in each_dict:
                if field_name in each_dict[key] and "MaxLength" in each_dict[key] and "BackColor" in each_dict[key]:
                    # Для дальнейшей сортировки полученных значений записываем значения из других полей (косвенный признаки):
                    # Значение в поле "BackColor" используется для сортировки по строкам
                    # Значение в поле "MaxLength" используется для сортировки по столбцам
                    raw_list.append([each_dict[key]["BackColor"], each_dict[key]["MaxLength"], each_dict[key][field_name]])
                else:
                    print("Структура JSON-файла не соответствует целевой.")
                    return
        # Сортируем получившийся список
        raw_list.sort(key=lambda val: val[1])  # Первая сортировка - по столбцам
        raw_list.sort(key=lambda val: val[0])  # Вторая сортировка - по строкам

        # Преобразуем список, оставляя только найденные значения полей
        raw_list = [elem[2] for elem in raw_list]

        # Разбиваем список значений на группы по 4 шт для вставки в excel
        ret = []
        for i in range(0, len(raw_list), chunk):
            ret.append(raw_list[i:i + chunk])
        return ret


if __name__ == '__main__':
    a = JSONToExcelParser()

    for i, arg in enumerate(sys.argv):
        if i == 0:  # Пропускаем название нашего модуля Python
            continue
        else:
            a.parse_json(arg)

    a.save_as_excel("myExcel")
